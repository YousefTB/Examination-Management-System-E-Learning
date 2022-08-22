# ----------------------
# Examiner Module
# ----------------------

from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
from tkcalendar import Calendar
from tktimepicker import AnalogPicker, AnalogThemes, SpinTimePickerModern, SpinTimePickerOld
from tktimepicker import constants
import datetime
import ExaminerWindb as db
import openpyxl
from openpyxl.styles import Font
from tkinter import filedialog as fd
import login_window


def examiner_module(examiner_ID):
    # Create The examiner window
    examiner = Tk()
    examiner.config(bg="Blue")

    # Change examiner window title
    examiner.title("Examiner main page")

    # Make window icon
    icon = PhotoImage(file="exam.icon.png")
    examiner.iconphoto(False, icon)

    # To make the window appear at the center of the screen

    def set_size(root, width, height):
        screen_width = root.winfo_screenwidth()
        secreem_height = root.winfo_screenheight()
        x = int((screen_width - width)/2)
        y = int((secreem_height-height)/2)
        root.geometry(f'{width}x{height}+{x}+{y}')

    # Set Dimensions
    set_size(examiner, 1200, 750)
    examiner.resizable(False, False)

    # Create main frame
    frame1 = Frame(examiner, width=300, height=750, bg="Blue")
    frame1.pack(side="left")

    # To import image
    img = ImageTk.PhotoImage(Image.open("mainframe.jpg"))

    # Create a Label Widget to display the text or Image
    label = Label(frame1, image=img)
    label.pack(side="left")

    # Create the second frame
    frame2 = Frame(examiner, width=900, height=750, bg="White")
    frame2.pack(side="right")

    # Button functions

    def personframe():
        clear_frame()
        personal_information()

    def examesFrame():
        clear_frame()
        examesMang_frame()

    def questFrame():
        clear_frame()
        questions_frame()

    def studentExamFrame():
        clear_frame()
        exames_students_frame()

    def logout():
        examiner.destroy()
        login_window.Login_win()

    # Create Buttons
    pesonal_infbtn = Button(frame1, text="Personal info", width=15, font=("Tahoma", 18),
                            height=1, borderwidth=0, bg="#FFFFFF", fg="Black", command=personframe).place(x=40, y=30)

    examebtn = Button(frame1, text="Exams", width=15, font=("Tahoma", 18),
                      height=1, borderwidth=0, bg="#FFFFFF", fg="Black", command=examesFrame).place(x=40, y=100)

    questionsbtn = Button(frame1, text="Questions", width=15, font=("Tahoma", 18),
                          height=1, borderwidth=0, bg="#FFFFFF", fg="Black", command=questFrame).place(x=40, y=170)

    studentExamesbtn = Button(frame1, text="Students exams", width=15, font=("Tahoma", 18),
                              height=1, borderwidth=0, bg="#FFFFFF", fg="Black", command=studentExamFrame).place(x=40, y=240)

    exitbtn = Button(frame1, text="Log out", width=15, font=("Tahoma", 18),
                     height=1, borderwidth=0, bg="#FF0099", fg="White", command=logout).place(x=40, y=650)

    # clear frame2 function

    def clear_frame():
        for widgets in frame2.winfo_children():
            widgets.destroy()
#############################################################################################################
    # Personal Informarion frame

    def personal_information():
        # Personal variables
        examiner_name = StringVar()
        examiner_id = StringVar()
        examiner_gen = StringVar()
        examiner_national_num = StringVar()
        examiner_phone = StringVar()
        examiner_email = StringVar()
        examiner_dep = StringVar()

        # To get personal info from Data base
        person_info = db.personInfo_examiner(examiner_ID)

        examiner_id.set(str(person_info[0]))
        examiner_name.set(person_info[1])
        examiner_national_num.set(person_info[2])
        examiner_gen.set(person_info[3])
        examiner_email.set(person_info[4])
        examiner_phone.set(person_info[5])
        examiner_dep.set(person_info[6])

        # Indicator labels
        examiner_name_lbl = Label(frame2, text="Name :",
                                  height=2, font=("Tahoma", 18, 'bold'), bg="White").place(x=15, y=20)

        examiner_name_var = Label(frame2, textvariable=examiner_name,
                                  height=2, font=("Tahoma", 16), bg="White").place(x=250, y=24)

        examiner_id_lbl = Label(frame2, text="ID :",
                                height=2, font=("Tahoma", 18, 'bold'), bg="White").place(x=15, y=70)

        examiner_id_var = Label(frame2, textvariable=examiner_id,
                                height=2, font=("Tahoma", 16), bg="White").place(x=250, y=74)

        examiner_gender_lbl = Label(frame2, text="Gender :",
                                    height=2, font=("Tahoma", 18, 'bold'), bg="White").place(x=15, y=130)

        examiner_gender_var = Label(frame2, textvariable=examiner_gen,
                                    height=2, font=("Tahoma", 16), bg="White").place(x=250, y=134)

        examiner_national_number_lbl = Label(frame2, text="National number :",
                                             height=2, font=("Tahoma", 18, 'bold'), bg="White").place(x=15, y=180)

        examiner_national_number_var = Label(frame2, textvariable=examiner_national_num,
                                             height=2, font=("Tahoma", 16), bg="White").place(x=250, y=184)

        examiner_phone_lbl = Label(frame2, text="Phone number :",
                                   height=2, font=("Tahoma", 18, 'bold'), bg="White").place(x=15, y=230)

        examiner_phone_var = Label(frame2, textvariable=examiner_phone,
                                   height=2, font=("Tahoma", 16), bg="White").place(x=250, y=234)

        examiner_email_lbl = Label(frame2, text="Email :",
                                   height=2, font=("Tahoma", 18, 'bold'), bg="White").place(x=15, y=280)

        examiner_email_var = Label(frame2, textvariable=examiner_email,
                                   height=2, font=("Tahoma", 16), bg="White").place(x=250, y=284)

        examiner_dep_lbl = Label(frame2, text="Department :",
                                 height=2, font=("Tahoma", 18, 'bold'), bg="White").place(x=15, y=330)

        examiner_dep_var = Label(frame2, textvariable=examiner_dep,
                                 height=2, font=("Tahoma", 16), bg="White").place(x=250, y=334)

#############################################################################################################
    def examesMang_frame():
        columns = ('exam_id', 'ex_name', 'beg_dat',
                   'beg_time', 'end_dat', 'end_time', 'total_deg', 'examiner_id')

        tree = ttk.Treeview(frame2, columns=columns,
                            show='headings', height=15)

        # define headings
        tree.heading('exam_id', text='Exam id')
        tree.heading('ex_name', text='Exam name')
        tree.heading('beg_dat', text='Begining date')
        tree.heading('beg_time', text='Begining time')
        tree.heading('total_deg', text='Total degree')
        tree.heading('end_dat', text='End date')
        tree.heading('end_time', text='End time')
        tree.heading('examiner_id', text='Examiner id')

        # Show exams data on the table
        def show_exams(tree):
            exams_data = db.get_exams(examiner_ID)
            tree.delete(*tree.get_children())
            for e in exams_data:
                tree.insert('', END, values=e)

        show_exams(tree)

        # Select function
        def item_selected(event):
            selected_item = tree.selection()
            item = tree.item(selected_item)
            record = item['values']
            exame_id.set(record[0])
            exame_name.set(record[1])
            exame_beg_date.set(record[2])
            exame_beg_time.set(record[3])
            b_t = (exame_beg_time.get()).split(':')
            time_picker1.set24Hrs(int(b_t[0]))
            time_picker1.setMins(int(b_t[1]))
            exame_end_date.set(record[4])
            exame_end_time.set(record[5])
            e_t = (exame_end_time.get()).split(':')
            time_picker2.set24Hrs(int(e_t[0]))
            time_picker2.setMins(int(e_t[1]))
            exame_total_deg.set(record[6])

        tree.bind('<<TreeviewSelect>>', item_selected)

        # Table place
        tree.place(x=50, y=10, height=230, width=800)

        # add a vertical scrollbar to the table
        scrollbar = ttk.Scrollbar(frame2, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.place(x=850, y=10, height=230)

        # add a horizontal scrollbar
        scrollbar3 = ttk.Scrollbar(
            frame2, orient=HORIZONTAL, command=tree.xview)
        tree.configure(xscrollcommand=scrollbar3.set)
        scrollbar3.place(x=50, y=240, height=20, width=800)

        # Add labels
        exID_lbl = Label(frame2, text="Exam id :",
                         font=("Tahoma", 15, "bold"), height=2, bg="White")
        exID_lbl.place(x=25, y=275)

        exName_lbl = Label(frame2, text="Exam name :",
                           font=("Tahoma", 15, "bold"), height=2, bg="White")
        exName_lbl.place(x=25, y=335)

        ex_beg_dat_lbl = Label(frame2, text="Calendar to set dates:",
                               font=("Tahoma", 15, "bold"), height=2, bg="White")
        ex_beg_dat_lbl.place(x=25, y=395)

        ex_beg_time_lbl = Label(frame2, text="Beggining time :",
                                font=("Tahoma", 15, "bold"), height=2, bg="White")
        ex_beg_time_lbl.place(x=300, y=395)

        ex_end_time_lbl = Label(frame2, text="End time :",
                                font=("Tahoma", 15, "bold"), height=2, bg="White")
        ex_end_time_lbl.place(x=500, y=395)

        ex_deg_lbl = Label(frame2, text="Total degree :",
                           font=("Tahoma", 15, "bold"), height=2, bg="White")
        ex_deg_lbl.place(x=670, y=395)

        # Variables decleration
        exame_id = StringVar()
        exame_name = StringVar()
        exame_beg_date = StringVar()
        exame_end_date = StringVar()
        exame_beg_time = StringVar()
        exame_end_time = StringVar()
        exame_total_deg = StringVar()

        # Entries decleration
        ex_id_var_label = Label(frame2, textvariable=exame_id,
                                font=("Tahoma", 14), height=2)
        ex_id_var_label.place(x=240, y=275)

        ex_name_entry = Entry(frame2, width=40, font=(
            "Tahoma", 14), textvariable=exame_name, borderwidth=2)
        ex_name_entry.place(x=240, y=350)

        ex_deg_entry = Entry(frame2, width=5, font=(
            "Tahoma", 14), textvariable=exame_total_deg, borderwidth=2)
        ex_deg_entry.place(x=670, y=450)

        # Entering date
        cal = Calendar(frame2, selectmode='day', year=2022, month=7,
                       day=23)

        cal.place(x=25, y=450)

        # time picker For begining time
        time_picker1 = SpinTimePickerModern(frame2)
        time_picker1.addAll(constants.HOURS24)
        time_picker1.configureAll(bg="#404040", height=1, fg="#ffffff", font=("Times", 20), hoverbg="#404040",
                                  hovercolor="#d73333", clickedbg="#2e2d2d", clickedcolor="#d73333")
        time_picker1.configure_separator(bg="#404040", fg="#ffffff")

        time_picker1.place(x=300, y=450)

        # time picker For end time
        time_picker2 = SpinTimePickerModern(frame2)
        time_picker2.addAll(constants.HOURS24)
        time_picker2.configureAll(bg="#404040", height=1, fg="#ffffff", font=("Times", 20), hoverbg="#404040",
                                  hovercolor="#d73333", clickedbg="#2e2d2d", clickedcolor="#d73333")
        time_picker2.configure_separator(bg="#404040", fg="#ffffff")

        time_picker2.place(x=500, y=450)

        exame_beg_date.set("1/10/22")
        exame_end_date.set("")
        # labels to show dates
        ex_beg_date_var_label = Label(frame2, textvariable=exame_beg_date,
                                      font=("Arial", 13), height=2, borderwidth=2, bg="White", fg="Blue")
        ex_beg_date_var_label.place(x=425, y=520)

        ex_end_date_var_label = Label(frame2, textvariable=exame_end_date,
                                      font=("Arial", 13), height=2, borderwidth=2, bg="White", fg="Blue")
        ex_end_date_var_label.place(x=425, y=565)

        # Set dates buttons functions

        def set_beg_date():
            if len(exame_end_date.get()) != 0:
                end_dat_list = (exame_end_date.get()).split('/')
                beg_dat_list = (cal.get_date()).split('/')

                # To compare dates
                # date in yyyy/mm/dd format
                beg_d = datetime.datetime(
                    int('20'+beg_dat_list[2]), int(beg_dat_list[0]), int(beg_dat_list[1]))
                end_d = datetime.datetime(
                    int('20'+end_dat_list[2]), int(end_dat_list[0]), int(end_dat_list[1]))

                if beg_d <= end_d:
                    exame_beg_date.set(cal.get_date())
                else:
                    messagebox.showerror(
                        "Failed process", "Enter correct dates please")
            else:
                exame_beg_date.set(cal.get_date())

            date = (exame_beg_date.get()).split('/')

            for i in range(2):
                if len(date[i]) < 2:
                    date[i] = '0' + date[i]
            exame_beg_date.set(date[0]+'/'+date[1]+'/'+date[2])

        def set_end_date():
            try:
                end_dat_list = (cal.get_date()).split('/')
                beg_dat_list = (exame_beg_date.get()).split('/')

                # To compare dates
                # date in yyyy/mm/dd format
                beg_d = datetime.datetime(
                    int('20'+beg_dat_list[2]), int(beg_dat_list[0]), int(beg_dat_list[1]))
                end_d = datetime.datetime(
                    int('20'+end_dat_list[2]), int(end_dat_list[0]), int(end_dat_list[1]))

                if beg_d <= end_d:
                    exame_end_date.set(cal.get_date())

                    date = (exame_end_date.get()).split('/')
                    for i in range(2):
                        if len(date[i]) < 2:
                            date[i] = '0' + date[i]
                    exame_end_date.set(date[0]+'/'+date[1]+'/'+date[2])
                else:
                    messagebox.showerror(
                        "Failed process", "Enter correct dates please")
            except:
                messagebox.showerror(
                    "Failed process", "Enter correct dates please")

        # Function to check validity of time
        def valid_time_date():
            beg_time_h = time_picker1.hours()
            beg_time_m = time_picker1.minutes()
            end_time_h = time_picker2.hours()
            end_time_m = time_picker2.minutes()

            if len(exame_end_date.get()) != 0:
                if (exame_end_date.get()) == (exame_beg_date.get()):
                    if end_time_h > beg_time_h:
                        return True
                    elif (end_time_h == beg_time_h) and (end_time_m > beg_time_m):
                        return True
                    else:
                        return FALSE
                else:
                    return True
            else:
                return False

        # Set dates buttons
        set_beg_date_btn = Button(frame2, text="Set begining date", width=14,
                                  height=1, font=("Arial", 13), borderwidth=0, bg="#CC0066", fg="White", command=set_beg_date)
        set_beg_date_btn.place(x=280, y=530)

        set_end_date_btn = Button(frame2, text="Set end date", width=14,
                                  height=1, font=("Arial", 13), borderwidth=0, bg="#CC0066", fg="White", command=set_end_date)
        set_end_date_btn.place(x=280, y=575)

        # Format time function
        def format_time():
            beg_h = time_picker1.hours24()
            end_h = time_picker2.hours24()
            beg_m = time_picker1.minutes()
            end_m = time_picker2.minutes()

            if beg_h < 10:
                beg_h = '0'+str(beg_h)
            else:
                beg_h = str(beg_h)

            if beg_m < 10:
                beg_m = '0'+str(beg_m)
            else:
                beg_m = str(beg_m)

            if end_h < 10:
                end_h = '0'+str(end_h)
            else:
                end_h = str(end_h)

            if end_m < 10:
                end_m = '0'+str(end_m)
            else:
                end_m = str(end_m)

            exame_beg_time.set(beg_h+':'+beg_m)
            exame_end_time.set(end_h+':'+end_m)

        # Clear function
        def clear():
            exame_end_date.set("")
            exame_name.set("")
            exame_total_deg.set("")
            exame_id.set("")

        # Buttons functions

        def insert_ex():
            try:
                if valid_time_date():
                    format_time()
                    db.insert_exams(exame_name.get(), exame_beg_date.get(), exame_beg_time.get(
                    ), exame_end_date.get(), exame_end_time.get(), float(exame_total_deg.get()), examiner_ID)
                    messagebox.showinfo("Successful process",
                                        "Inserted successfully")
                    show_exams(tree)
                    clear()
                else:
                    messagebox.showerror(
                        "Failed process", "Enter correct information please")
            except:
                messagebox.showerror(
                    "Failed process", "Invalid information")

        def update_ex():
            try:
                if valid_time_date():
                    format_time()
                    db.update_exams(int(exame_id.get()), exame_name.get(), exame_beg_date.get(), exame_beg_time.get(
                    ), exame_end_date.get(), exame_end_time.get(), float(exame_total_deg.get()), examiner_ID)
                    messagebox.showinfo("Successful process",
                                        "Updated successfully")
                    show_exams(tree)
                    clear()
                else:
                    messagebox.showerror(
                        "Failed process", "Enter correct information please")
            except:
                messagebox.showerror(
                    "Failed process", "Invalid information")

        def delete_ex():
            try:
                db.delete_exams(int(exame_id.get()))
                messagebox.showinfo("Successful process",
                                    "Deleted successfully")
                show_exams(tree)
                clear()
            except:
                messagebox.showerror(
                    "Failed process", "Invalid information")

        # Add function buttons
        insert_ex_btn = Button(frame2, text="Insert", width=12,
                               height=1, font=("Arial", 14), borderwidth=0, bg="#FF0099", fg="White", command=insert_ex)
        insert_ex_btn.place(x=400, y=700)

        update_ex_btn = Button(frame2, text="Update", width=12,
                               height=1, font=("Arial", 14), borderwidth=0, bg="#FF0099", fg="White", command=update_ex)
        update_ex_btn.place(x=550, y=700)

        delete_ex_btn = Button(frame2, text="Delete", width=12,
                               height=1, font=("Arial", 14), borderwidth=0, bg="#FF0099", fg="White", command=delete_ex)
        delete_ex_btn.place(x=700, y=700)

###################################################################################################

    def questions_frame():
        columns = ('q_id', 'q_name', 'ans1',
                   'ans2', 'ans3', 'ans4', 'answer', 'deg', 'exam_id')

        tree3 = ttk.Treeview(frame2, columns=columns,
                             show='headings', height=15)

        # define headings
        tree3.heading('q_id', text='ID')
        tree3.heading('q_name', text='Question')
        tree3.heading('ans1', text='1st answer')
        tree3.heading('ans2', text='2nd answer')
        tree3.heading('ans3', text='3rd answer')
        tree3.heading('ans4', text='4th answer')
        tree3.heading('answer', text='Answer')
        tree3.heading('deg', text='Degree')
        tree3.heading('exam_id', text='Exam id')

        # To show questions data with particular exam
        def show_ques_with_particular_exam(tree3, exam_id):
            question_data = db.get_question(exam_id)
            tree3.delete(*tree3.get_children())
            for q in question_data:
                tree3.insert('', END, values=q)

        # Select function
        def item_selected(event):
            selected_item = tree3.selection()
            item = tree3.item(selected_item)
            record = item['values']
            question_id.set(record[0])
            question.set(record[1])
            ans1.set(record[2])
            ans2.set(record[3])
            ans3.set(record[4])
            ans4.set(record[5])
            answer.set(record[6])
            question_deg.set(record[7])

        tree3.bind('<<TreeviewSelect>>', item_selected)

        # Table place
        tree3.place(x=50, y=10, height=230, width=800)

        # add a vertical scrollbar to the table
        scrollbar4 = ttk.Scrollbar(
            frame2, orient=VERTICAL, command=tree3.yview)
        tree3.configure(yscroll=scrollbar4.set)
        scrollbar4.place(x=850, y=10, height=240)

        # add a horizontal scrollbar
        scrollbar5 = ttk.Scrollbar(
            frame2, orient=HORIZONTAL, command=tree3.xview)
        tree3.configure(xscrollcommand=scrollbar5.set)
        scrollbar5.place(x=50, y=240, height=20, width=800)

        # Add labels
        quesID_lbl = Label(frame2, text="Question id :",
                           font=("Tahoma", 15, "bold"), height=2, bg="White")
        quesID_lbl.place(x=25, y=275)

        ques_deg_lbl = Label(frame2, text="Degree :",
                             font=("Tahoma", 15, "bold"), height=2, bg="White")
        ques_deg_lbl.place(x=400, y=275)

        ques_lbl = Label(frame2, text="question*",
                         font=("Tahoma", 13), height=2, bg="White")
        ques_lbl.place(x=485, y=310)

        ans_lbl = Label(frame2, text="The answer is:",
                        font=("Tahoma", 15, "bold"), height=2, bg="White")
        ans_lbl.place(x=25, y=440)

        ex_lbl = Label(frame2, text="The Exam :",
                       font=("Tahoma", 15, "bold"), height=2, bg="White")
        ex_lbl.place(x=25, y=490)

        # Variables that hold student values
        question_id = StringVar()
        question = StringVar()
        ans1 = StringVar()
        ans2 = StringVar()
        ans3 = StringVar()
        ans4 = StringVar()
        answer = StringVar()
        question_deg = StringVar()
        exam = StringVar()

        # places to show and write students data
        question_id_var_label = Label(frame2, textvariable=question_id,
                                      font=("Tahoma", 14), height=2, bg="White")
        question_id_var_label.place(x=180, y=280)

        ques_entry = Entry(frame2, width=40, font=(
            "Tahoma", 15), textvariable=question, border=2)
        ques_entry.place(x=25, y=325)

        deg_entry = Entry(frame2, width=5, font=(
            "Tahoma", 15), textvariable=question_deg, border=2)
        deg_entry.place(x=500, y=285)

        ans1_entry = Entry(frame2, width=15, font=(
            "Tahoma", 15), textvariable=ans1, border=1)
        ans1_entry.place(x=30, y=365)

        ans2_entry = Entry(frame2, width=15, font=(
            "Tahoma", 15), textvariable=ans2, border=1)
        ans2_entry.place(x=250, y=365)

        ans3_entry = Entry(frame2, width=15, font=(
            "Tahoma", 15), textvariable=ans3, border=1)
        ans3_entry.place(x=480, y=365)

        ans4_entry = Entry(frame2, width=15, font=(
            "Tahoma", 15), textvariable=ans4, border=1)
        ans4_entry.place(x=710, y=365)

        answer_var_lbl = Label(frame2, textvariable=answer,
                               font=("Tahoma", 14), height=2, bg="White")
        answer_var_lbl.place(x=230, y=440)

        def selected_1():
            answer.set(ans1.get())

        def selected_2():
            answer.set(ans2.get())

        def selected_3():
            answer.set(ans3.get())

        def selected_4():
            answer.set(ans4.get())

        s = IntVar()
        ans1_rb = Radiobutton(frame2, text="1st", variable=s,
                              value=1, font=("Arial", 14), command=selected_1, bg="White")
        ans1_rb.place(x=30, y=390)

        ans2_rb = Radiobutton(frame2, text="2nd", variable=s,
                              value=2, font=("Arial", 14), command=selected_2, bg="White")
        ans2_rb.place(x=260, y=390)

        ans3_rb = Radiobutton(frame2, text="3rd", variable=s,
                              value=3, font=("Arial", 14), command=selected_3, bg="White")
        ans3_rb.place(x=490, y=390)

        ans4_rb = Radiobutton(frame2, text="4th", variable=s,
                              value=4, font=("Arial", 14), command=selected_4, bg="White")
        ans4_rb.place(x=720, y=390)

        ex_cbx = ttk.Combobox(frame2, width=40, textvariable=exam, font=(
            "Tahoma", 13), background="White")

        ex_in_list = []
        ex_in_dic = {}
        db.get_exams_combo(ex_in_list, ex_in_dic, examiner_ID)
        ex_cbx['values'] = ex_in_list
        ex_cbx['state'] = 'readonly'
        ex_cbx.bind("<<ComboboxSelected>>", lambda _: show_ques_with_particular_exam(
            tree3, ex_in_dic[exam.get()]))
        ex_cbx.place(x=230, y=500)

        # Clear function
        def clear():
            question_id.set("")
            question.set("")
            ans1.set("")
            ans2.set("")
            ans3.set("")
            ans4.set("")
            answer.set("")

        # Buttons functions

        def insert_question():
            try:
                db.insert_question(question.get(), ans1.get(), ans2.get(), ans3.get(
                ), ans4.get(), answer.get(), question_deg.get(), ex_in_dic[exam.get()])
                messagebox.showinfo("Successful process",
                                    "Inserted successfully")
                show_ques_with_particular_exam(
                    tree3, ex_in_dic[exam.get()])
                clear()
            except:
                messagebox.showerror(
                    "Failed process", "Please try again with correct data")

        def update_question():
            try:
                db.update_question(int(question_id.get()), question.get(), ans1.get(), ans2.get(), ans3.get(
                ), ans4.get(), answer.get(), question_deg.get(), ex_in_dic[exam.get()])
                messagebox.showinfo("Successful process",
                                    "Update successfully")
                show_ques_with_particular_exam(
                    tree3, ex_in_dic[exam.get()])
                clear()
            except:
                messagebox.showerror(
                    "Failed process", "Please try again with correct data")

        def delete_question():
            try:
                db.delete_question(int(question_id.get()))
                messagebox.showinfo("Successful process",
                                    "Deleted successfully")
                show_ques_with_particular_exam(
                    tree3, ex_in_dic[exam.get()])
                clear()
            except:
                messagebox.showerror(
                    "Failed process", "Please try again with correct data")

        def get_sum_degs():
            try:
                messagebox.showinfo("The total degree",
                                    f"The sum is :{db.get_total_question_degrees(ex_in_dic[exam.get()])}")
            except:
                messagebox.showerror(
                    "Failed process", "Please choose an exam")

        # Add function buttons
        insert_question_btn = Button(frame2, text="Insert", width=12,
                                     height=1, font=("Arial", 14), borderwidth=0, bg="#FF0099", fg="White", command=insert_question)
        insert_question_btn.place(x=400, y=700)

        update_question_btn = Button(frame2, text="Update", width=12,
                                     height=1, font=("Arial", 14), borderwidth=0, bg="#FF0099", fg="White", command=update_question)
        update_question_btn.place(x=550, y=700)

        delete_question_btn = Button(frame2, text="Delete", width=12,
                                     height=1, font=("Arial", 14), borderwidth=0, bg="#FF0099", fg="White", command=delete_question)
        delete_question_btn.place(x=700, y=700)

        get_ques_deg_btn = Button(frame2, text="Get sum of question degrees", width=23,
                                  height=1, font=("Arial", 14), borderwidth=0, bg="#CC0066", fg="White", command=get_sum_degs)
        get_ques_deg_btn.place(x=30, y=550)

####################################################################################################

    def exames_students_frame():
        #  Exam id variable definition
        exam = StringVar()
        stu_id = StringVar()

        columns = ('ex_id', 'ex_name', 'stu_id',
                   'stu_name', 'status', 'stu_deg')

        tree = ttk.Treeview(frame2, columns=columns,
                            show='headings', height=15)

        # define headings
        tree.heading('ex_id', text='Exam id')
        tree.heading('ex_name', text='Exam name')
        tree.heading('stu_id', text='Student id')
        tree.heading('stu_name', text='Student name')
        tree.heading('status', text='Status')
        tree.heading('stu_deg', text='Student degree')

        ################ To show student exams data #####################
        def show_students_with_exams(tree, exam_ID):
            students_data = db.get_students_exams(exam_ID)
            tree.delete(*tree.get_children())
            for s in students_data:
                tree.insert('', END, values=s)
        # Table place
        tree.place(x=50, y=10, height=230, width=800)

        # Select function
        def item_selected(event):
            selected_item = tree.selection()
            item = tree.item(selected_item)
            record = item['values']
            exam.set(record[1])
            stu_id.set(record[2])

        tree.bind('<<TreeviewSelect>>', item_selected)

        # add a vertical scrollbar to the table
        scrollbar = ttk.Scrollbar(frame2, orient=VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.place(x=850, y=10, height=230)

        # add a horizontal scrollbar
        scrollbar3 = ttk.Scrollbar(
            frame2, orient=HORIZONTAL, command=tree.xview)
        tree.configure(xscrollcommand=scrollbar3.set)
        scrollbar3.place(x=50, y=240, height=20, width=800)

        # Add labels
        stu_lbl = Label(frame2, text="Students", font=(
            "Tahoma", 17, "bold"), height=2, bg="White")
        stu_lbl.place(x=50, y=260)

        exames_lbl = Label(frame2, text="Exams", font=(
            "Tahoma", 17, "bold"), height=2, bg="White")
        exames_lbl.place(x=420, y=260)

        stu_id_lbl = Label(frame2, text="Student id:", font=(
            "Tahoma", 15, "bold"), height=2, bg="White")
        stu_id_lbl.place(x=25, y=580)

        stu_id_var_lbl = Label(frame2, textvariable=stu_id, font=(
            "Tahoma", 14), height=2, bg="White")
        stu_id_var_lbl.place(x=150, y=580)

        # Exames combobox
        ex_cbx = ttk.Combobox(frame2, width=30, textvariable=exam, font=(
            "Tahoma", 13), background="White")
        ex_in_list = []
        ex_in_dic = {}
        db.get_exams_combo(ex_in_list, ex_in_dic, examiner_ID)
        ex_cbx['values'] = ex_in_list
        ex_cbx['state'] = 'readonly'
        ex_cbx.bind("<<ComboboxSelected>>", lambda _: show_students_with_exams(
            tree, ex_in_dic[exam.get()]))

        ex_cbx.place(x=420, y=330)

        ########################## Another table ############################
        columns2 = ('stu_id', 'stu_name')

        tree2 = ttk.Treeview(frame2, columns=columns2,
                             show='headings', height=15)

        # define headings
        tree2.heading('stu_id', text='Student id')
        tree2.heading('stu_name', text='Student name')

        # To show students data
        def show_students(tree2, examiner_ID):
            students_data = db.get_students_v2(examiner_ID)
            tree2.delete(*tree2.get_children())
            for s in students_data:
                tree2.insert('', END, values=s)

        show_students(tree2, examiner_ID)

        # Table place
        tree2.place(x=50, y=330, height=230, width=300)

        # Select function
        def item_selected(event):
            selected_item = tree2.selection()
            item = tree2.item(selected_item)
            record = item['values']
            stu_id.set(record[0])

        tree2.bind('<<TreeviewSelect>>', item_selected)

        # add a vertical scrollbar to the table
        scrollbar1 = ttk.Scrollbar(
            frame2, orient=VERTICAL, command=tree2.yview)
        tree2.configure(yscroll=scrollbar1.set)
        scrollbar1.place(x=350, y=330, height=230)

        # add a horizontal scrollbar
        scrollbar2 = ttk.Scrollbar(
            frame2, orient=HORIZONTAL, command=tree2.xview)
        tree2.configure(xscrollcommand=scrollbar2.set)
        scrollbar2.place(x=50, y=560, height=20, width=300)

        # Buttons functions

        def insert():
            try:
                db.insert_exam_to_student(
                    int(stu_id.get()), ex_in_dic[exam.get()])
                q_list = db.get_questions_ids(ex_in_dic[exam.get()])
                for q in q_list:
                    db.insert_question_to_student(int(stu_id.get()), q)
                messagebox.showinfo("Successful process",
                                    "Inserted successfully")
                show_students_with_exams(tree, ex_in_dic[exam.get()])
                stu_id.set("")
            except:
                messagebox.showerror("Failed process", "Please try again")

        def delete():
            try:
                db.delete_exam_to_student(
                    int(stu_id.get()), ex_in_dic[exam.get()])
                q_list = db.get_questions_ids(ex_in_dic[exam.get()])
                for q in q_list:
                    db.delete_question_to_student(int(stu_id.get()), q)
                messagebox.showinfo("Successful process",
                                    "Deleted successfully")
                show_students_with_exams(tree, ex_in_dic[exam.get()])
                stu_id.set("")
            except:
                messagebox.showerror("Failed process", "Please try again")

        # Add function buttons
        insert_question_btn = Button(frame2, text="Insert", width=12,
                                     height=1, font=("Arial", 14), borderwidth=0, bg="#FF0099", fg="White", command=insert)
        insert_question_btn.place(x=550, y=700)

        delete_question_btn = Button(frame2, text="Delete", width=12,
                                     height=1, font=("Arial", 14), borderwidth=0, bg="#FF0099", fg="White", command=delete)
        delete_question_btn.place(x=700, y=700)

        # Exel sheet function

        def get_exel_sheet():
            try:
                # To bring data
                data = db.get_students_exams_for_exel(
                    int(ex_in_dic[exam.get()]))

                # indicate the file location
                filenames = fd.askdirectory(
                    title='Choose location',
                    initialdir='/')

                location = filenames + '/' + exam.get() + ' marks.xlsx'

                # Call a Workbook() function of openpyxl
                # to create a new blank Workbook object
                wb = openpyxl.Workbook()

                # Get workbook active sheet
                # from the active attribute.
                sheet = wb.active

                # set columns width to fit data
                sheet.column_dimensions['A'].width = 20
                sheet.column_dimensions['B'].width = 25
                sheet.column_dimensions['C'].width = 20

                # Font style for heading
                sheet.cell(row=1, column=1).font = Font(
                    size=16, bold=True, name='Times New Roman')
                sheet.cell(row=1, column=1).value = "Student ID"

                sheet.cell(row=1, column=2).font = Font(
                    size=16, bold=True, name='Times New Roman')
                sheet.cell(row=1, column=2).value = "Student name"

                sheet.cell(row=1, column=3).font = Font(
                    size=16, bold=True, name='Times New Roman')
                sheet.cell(row=1, column=3).value = "Student mark"

                for row in data:
                    sheet.append(row)

                wb.save(location)
                messagebox.showinfo("Successful process",
                                    "Saved successfully")
            except:
                messagebox.showerror("Failed process",
                                     "Please choose an exam")

        # Add button for using exel sheet
        exel_btn = Button(frame2, text="Get degrees in excel sheet", width=30,
                          height=1, font=("Arial", 14), borderwidth=0, bg="#197A19", fg="White", command=get_exel_sheet)
        exel_btn.place(x=420, y=545)

    # Run app infinitely
    examiner.mainloop()


# examiner_module(1)
